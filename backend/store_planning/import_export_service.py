"""
开店计划数据导入导出服务
"""
import pandas as pd
import io
from typing import Dict, List, Any, Tuple
from decimal import Decimal, InvalidOperation
from datetime import datetime, date
from django.db import transaction, models
from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model
from .models import StorePlan, RegionalPlan, BusinessRegion, StoreType, PlanExecutionLog

User = get_user_model()


class PlanImportExportService:
    """计划导入导出服务类"""
    
    # Excel列名映射
    IMPORT_COLUMNS = {
        'plan_name': '计划名称',
        'plan_type': '计划类型',
        'start_date': '开始日期',
        'end_date': '结束日期',
        'description': '计划描述',
        'region_name': '经营区域',
        'region_code': '区域编码',
        'store_type_name': '门店类型',
        'store_type_code': '类型编码',
        'target_count': '目标数量',
        'contribution_rate': '贡献率(%)',
        'budget_amount': '预算金额'
    }
    
    # 计划类型映射
    PLAN_TYPE_MAPPING = {
        '年度计划': 'annual',
        '季度计划': 'quarterly',
        'annual': 'annual',
        'quarterly': 'quarterly'
    }
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.error_count = 0
    
    def validate_excel_structure(self, df: pd.DataFrame) -> bool:
        """验证Excel文件结构"""
        required_columns = list(self.IMPORT_COLUMNS.values())
        missing_columns = []
        
        for col in required_columns:
            if col not in df.columns:
                missing_columns.append(col)
        
        if missing_columns:
            self.errors.append({
                'type': 'structure_error',
                'message': f'缺少必需的列: {", ".join(missing_columns)}',
                'row': None
            })
            return False
        
        return True
    
    def clean_and_validate_row_data(self, row: pd.Series, row_index: int) -> Dict[str, Any]:
        """清理和验证行数据"""
        cleaned_data = {}
        row_errors = []
        
        try:
            # 计划名称
            plan_name = str(row['计划名称']).strip() if pd.notna(row['计划名称']) else ''
            if not plan_name:
                row_errors.append('计划名称不能为空')
            else:
                cleaned_data['plan_name'] = plan_name
            
            # 计划类型
            plan_type_raw = str(row['计划类型']).strip() if pd.notna(row['计划类型']) else ''
            if not plan_type_raw:
                row_errors.append('计划类型不能为空')
            elif plan_type_raw not in self.PLAN_TYPE_MAPPING:
                row_errors.append(f'计划类型无效: {plan_type_raw}，支持的类型: {", ".join(self.PLAN_TYPE_MAPPING.keys())}')
            else:
                cleaned_data['plan_type'] = self.PLAN_TYPE_MAPPING[plan_type_raw]
            
            # 开始日期
            try:
                if pd.notna(row['开始日期']):
                    if isinstance(row['开始日期'], str):
                        start_date = pd.to_datetime(row['开始日期']).date()
                    else:
                        start_date = row['开始日期'].date() if hasattr(row['开始日期'], 'date') else row['开始日期']
                    cleaned_data['start_date'] = start_date
                else:
                    row_errors.append('开始日期不能为空')
            except (ValueError, TypeError):
                row_errors.append(f'开始日期格式无效: {row["开始日期"]}')
            
            # 结束日期
            try:
                if pd.notna(row['结束日期']):
                    if isinstance(row['结束日期'], str):
                        end_date = pd.to_datetime(row['结束日期']).date()
                    else:
                        end_date = row['结束日期'].date() if hasattr(row['结束日期'], 'date') else row['结束日期']
                    cleaned_data['end_date'] = end_date
                else:
                    row_errors.append('结束日期不能为空')
            except (ValueError, TypeError):
                row_errors.append(f'结束日期格式无效: {row["结束日期"]}')
            
            # 验证日期范围
            if 'start_date' in cleaned_data and 'end_date' in cleaned_data:
                if cleaned_data['start_date'] >= cleaned_data['end_date']:
                    row_errors.append('结束日期必须晚于开始日期')
            
            # 计划描述（可选）
            description = str(row['计划描述']).strip() if pd.notna(row['计划描述']) else ''
            cleaned_data['description'] = description
            
            # 经营区域
            region_name = str(row['经营区域']).strip() if pd.notna(row['经营区域']) else ''
            region_code = str(row['区域编码']).strip() if pd.notna(row['区域编码']) else ''
            
            if not region_name and not region_code:
                row_errors.append('经营区域名称或编码至少需要提供一个')
            else:
                # 查找区域
                region = None
                if region_code:
                    try:
                        region = BusinessRegion.objects.get(code=region_code, is_active=True)
                    except BusinessRegion.DoesNotExist:
                        pass
                
                if not region and region_name:
                    try:
                        region = BusinessRegion.objects.get(name=region_name, is_active=True)
                    except BusinessRegion.DoesNotExist:
                        pass
                
                if not region:
                    row_errors.append(f'找不到经营区域: {region_name or region_code}')
                else:
                    cleaned_data['region'] = region
            
            # 门店类型
            store_type_name = str(row['门店类型']).strip() if pd.notna(row['门店类型']) else ''
            store_type_code = str(row['类型编码']).strip() if pd.notna(row['类型编码']) else ''
            
            if not store_type_name and not store_type_code:
                row_errors.append('门店类型名称或编码至少需要提供一个')
            else:
                # 查找门店类型
                store_type = None
                if store_type_code:
                    try:
                        store_type = StoreType.objects.get(code=store_type_code, is_active=True)
                    except StoreType.DoesNotExist:
                        pass
                
                if not store_type and store_type_name:
                    try:
                        store_type = StoreType.objects.get(name=store_type_name, is_active=True)
                    except StoreType.DoesNotExist:
                        pass
                
                if not store_type:
                    row_errors.append(f'找不到门店类型: {store_type_name or store_type_code}')
                else:
                    cleaned_data['store_type'] = store_type
            
            # 目标数量
            try:
                if pd.notna(row['目标数量']):
                    target_count = int(float(row['目标数量']))
                    if target_count <= 0:
                        row_errors.append('目标数量必须大于0')
                    else:
                        cleaned_data['target_count'] = target_count
                else:
                    row_errors.append('目标数量不能为空')
            except (ValueError, TypeError):
                row_errors.append(f'目标数量格式无效: {row["目标数量"]}')
            
            # 贡献率（可选）
            if pd.notna(row['贡献率(%)']):
                try:
                    contribution_rate = Decimal(str(row['贡献率(%)']))
                    if contribution_rate < 0 or contribution_rate > 100:
                        row_errors.append('贡献率必须在0-100之间')
                    else:
                        cleaned_data['contribution_rate'] = contribution_rate
                except (ValueError, InvalidOperation):
                    row_errors.append(f'贡献率格式无效: {row["贡献率(%)"]}')
            
            # 预算金额（可选）
            if pd.notna(row['预算金额']):
                try:
                    budget_amount = Decimal(str(row['预算金额']))
                    if budget_amount < 0:
                        row_errors.append('预算金额不能为负数')
                    else:
                        cleaned_data['budget_amount'] = budget_amount
                except (ValueError, InvalidOperation):
                    row_errors.append(f'预算金额格式无效: {row["预算金额"]}')
            else:
                cleaned_data['budget_amount'] = Decimal('0.00')
        
        except Exception as e:
            row_errors.append(f'数据处理异常: {str(e)}')
        
        # 记录行错误
        if row_errors:
            for error in row_errors:
                self.errors.append({
                    'type': 'data_error',
                    'message': error,
                    'row': row_index + 2  # Excel行号从2开始（第1行是标题）
                })
        
        return cleaned_data if not row_errors else None
    
    def group_data_by_plan(self, validated_data: List[Dict[str, Any]]) -> Dict[str, Dict]:
        """按计划分组数据"""
        plans_data = {}
        
        for row_data in validated_data:
            plan_key = (
                row_data['plan_name'],
                row_data['plan_type'],
                row_data['start_date'],
                row_data['end_date']
            )
            
            if plan_key not in plans_data:
                plans_data[plan_key] = {
                    'plan_info': {
                        'name': row_data['plan_name'],
                        'plan_type': row_data['plan_type'],
                        'start_date': row_data['start_date'],
                        'end_date': row_data['end_date'],
                        'description': row_data['description']
                    },
                    'regional_plans': []
                }
            
            # 检查区域计划重复
            region_store_key = (row_data['region'].id, row_data['store_type'].id)
            existing_regional = [
                rp for rp in plans_data[plan_key]['regional_plans']
                if (rp['region'].id, rp['store_type'].id) == region_store_key
            ]
            
            if existing_regional:
                self.warnings.append({
                    'type': 'duplicate_warning',
                    'message': f'计划"{row_data["plan_name"]}"中区域"{row_data["region"].name}"和门店类型"{row_data["store_type"].name}"的组合重复，将使用最后一条记录'
                })
                # 移除重复的记录
                plans_data[plan_key]['regional_plans'] = [
                    rp for rp in plans_data[plan_key]['regional_plans']
                    if (rp['region'].id, rp['store_type'].id) != region_store_key
                ]
            
            plans_data[plan_key]['regional_plans'].append({
                'region': row_data['region'],
                'store_type': row_data['store_type'],
                'target_count': row_data['target_count'],
                'contribution_rate': row_data.get('contribution_rate'),
                'budget_amount': row_data.get('budget_amount', Decimal('0.00'))
            })
        
        return plans_data
    
    def validate_plan_business_rules(self, plans_data: Dict[str, Dict]) -> bool:
        """验证计划业务规则"""
        valid = True
        
        for plan_key, plan_data in plans_data.items():
            plan_name = plan_data['plan_info']['name']
            regional_plans = plan_data['regional_plans']
            
            # 验证总贡献率不超过100%
            total_contribution_rate = sum(
                rp.get('contribution_rate', Decimal('0.00')) or Decimal('0.00')
                for rp in regional_plans
            )
            
            if total_contribution_rate > 100:
                self.errors.append({
                    'type': 'business_rule_error',
                    'message': f'计划"{plan_name}"的总贡献率({total_contribution_rate}%)超过100%',
                    'row': None
                })
                valid = False
            
            # 检查计划名称是否已存在
            if StorePlan.objects.filter(name=plan_name).exists():
                self.errors.append({
                    'type': 'business_rule_error',
                    'message': f'计划名称"{plan_name}"已存在',
                    'row': None
                })
                valid = False
        
        return valid
    
    @transaction.atomic
    def import_plans_from_excel(self, file_content: bytes, user: User) -> Dict[str, Any]:
        """从Excel导入计划数据"""
        self.errors = []
        self.warnings = []
        self.success_count = 0
        self.error_count = 0
        
        try:
            # 读取Excel文件
            df = pd.read_excel(io.BytesIO(file_content))
            
            # 验证文件结构
            if not self.validate_excel_structure(df):
                return self._generate_import_result()
            
            # 清理和验证每行数据
            validated_data = []
            for index, row in df.iterrows():
                cleaned_data = self.clean_and_validate_row_data(row, index)
                if cleaned_data:
                    validated_data.append(cleaned_data)
                else:
                    self.error_count += 1
            
            # 如果有数据错误，停止导入
            if self.errors:
                return self._generate_import_result()
            
            # 按计划分组数据
            plans_data = self.group_data_by_plan(validated_data)
            
            # 验证业务规则
            if not self.validate_plan_business_rules(plans_data):
                return self._generate_import_result()
            
            # 创建计划
            created_plans = []
            for plan_key, plan_data in plans_data.items():
                try:
                    plan = self._create_plan_with_regional_plans(
                        plan_data['plan_info'],
                        plan_data['regional_plans'],
                        user
                    )
                    created_plans.append(plan)
                    self.success_count += 1
                except Exception as e:
                    self.errors.append({
                        'type': 'creation_error',
                        'message': f'创建计划"{plan_data["plan_info"]["name"]}"失败: {str(e)}',
                        'row': None
                    })
                    self.error_count += 1
            
            return self._generate_import_result(created_plans)
        
        except Exception as e:
            self.errors.append({
                'type': 'file_error',
                'message': f'文件处理失败: {str(e)}',
                'row': None
            })
            return self._generate_import_result()
    
    def _create_plan_with_regional_plans(self, plan_info: Dict, regional_plans: List[Dict], user: User) -> StorePlan:
        """创建计划及其区域计划"""
        # 创建主计划
        plan = StorePlan.objects.create(
            name=plan_info['name'],
            plan_type=plan_info['plan_type'],
            start_date=plan_info['start_date'],
            end_date=plan_info['end_date'],
            description=plan_info['description'],
            created_by=user
        )
        
        # 创建区域计划
        total_target_count = 0
        total_budget_amount = Decimal('0.00')
        
        for regional_plan_data in regional_plans:
            RegionalPlan.objects.create(
                plan=plan,
                region=regional_plan_data['region'],
                store_type=regional_plan_data['store_type'],
                target_count=regional_plan_data['target_count'],
                contribution_rate=regional_plan_data.get('contribution_rate'),
                budget_amount=regional_plan_data['budget_amount']
            )
            total_target_count += regional_plan_data['target_count']
            total_budget_amount += regional_plan_data['budget_amount']
        
        # 更新计划汇总数据
        plan.total_target_count = total_target_count
        plan.total_budget_amount = total_budget_amount
        plan.save(update_fields=['total_target_count', 'total_budget_amount'])
        
        # 记录执行日志
        PlanExecutionLog.objects.create(
            plan=plan,
            action_type='plan_created',
            action_description=f'通过Excel导入创建计划：{plan.name}',
            created_by=user
        )
        
        return plan
    
    def _generate_import_result(self, created_plans: List[StorePlan] = None) -> Dict[str, Any]:
        """生成导入结果"""
        return {
            'success': len(self.errors) == 0,
            'total_processed': self.success_count + self.error_count,
            'success_count': self.success_count,
            'error_count': self.error_count,
            'created_plans': [
                {
                    'id': plan.id,
                    'name': plan.name,
                    'plan_type': plan.get_plan_type_display(),
                    'total_target_count': plan.total_target_count
                }
                for plan in (created_plans or [])
            ],
            'errors': self.errors,
            'warnings': self.warnings
        }
    
    def generate_import_template(self) -> bytes:
        """生成导入模板Excel文件"""
        # 创建示例数据
        sample_data = [
            {
                '计划名称': '2024年华东区开店计划',
                '计划类型': '年度计划',
                '开始日期': '2024-01-01',
                '结束日期': '2024-12-31',
                '计划描述': '2024年华东区域开店计划',
                '经营区域': '华东区',
                '区域编码': 'HD',
                '门店类型': '直营店',
                '类型编码': 'ZY',
                '目标数量': 50,
                '贡献率(%)': 30.5,
                '预算金额': 5000000
            },
            {
                '计划名称': '2024年华东区开店计划',
                '计划类型': '年度计划',
                '开始日期': '2024-01-01',
                '结束日期': '2024-12-31',
                '计划描述': '2024年华东区域开店计划',
                '经营区域': '华东区',
                '区域编码': 'HD',
                '门店类型': '加盟店',
                '类型编码': 'JM',
                '目标数量': 30,
                '贡献率(%)': 20.0,
                '预算金额': 3000000
            }
        ]
        
        # 创建DataFrame
        df = pd.DataFrame(sample_data)
        
        # 生成Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='开店计划导入模板', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['开店计划导入模板']
            
            # 设置列宽
            column_widths = {
                'A': 20,  # 计划名称
                'B': 12,  # 计划类型
                'C': 12,  # 开始日期
                'D': 12,  # 结束日期
                'E': 25,  # 计划描述
                'F': 15,  # 经营区域
                'G': 12,  # 区域编码
                'H': 15,  # 门店类型
                'I': 12,  # 类型编码
                'J': 12,  # 目标数量
                'K': 12,  # 贡献率
                'L': 15   # 预算金额
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        output.seek(0)
        return output.getvalue() 
   
    def export_plans_to_excel(self, plan_ids=None, start_date=None, end_date=None, 
                             plan_type=None, status=None) -> bytes:
        """导出计划数据到Excel"""
        # 构建查询条件
        queryset = StorePlan.objects.select_related('created_by').prefetch_related(
            'regional_plans__region',
            'regional_plans__store_type'
        )
        
        # 应用过滤条件
        if plan_ids:
            queryset = queryset.filter(id__in=plan_ids)
        
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        if plan_type:
            queryset = queryset.filter(plan_type=plan_type)
        
        if status:
            queryset = queryset.filter(status=status)
        
        # 准备导出数据
        export_data = []
        
        for plan in queryset:
            for regional_plan in plan.regional_plans.all():
                export_data.append({
                    '计划名称': plan.name,
                    '计划类型': plan.get_plan_type_display(),
                    '计划状态': plan.get_status_display(),
                    '开始日期': plan.start_date.strftime('%Y-%m-%d'),
                    '结束日期': plan.end_date.strftime('%Y-%m-%d'),
                    '计划描述': plan.description or '',
                    '经营区域': regional_plan.region.name,
                    '区域编码': regional_plan.region.code,
                    '门店类型': regional_plan.store_type.name,
                    '类型编码': regional_plan.store_type.code,
                    '目标数量': regional_plan.target_count,
                    '完成数量': regional_plan.completed_count,
                    '完成率(%)': regional_plan.completion_rate,
                    '贡献率(%)': float(regional_plan.contribution_rate) if regional_plan.contribution_rate else '',
                    '预算金额': float(regional_plan.budget_amount),
                    '创建人': plan.created_by.get_full_name() or plan.created_by.username,
                    '创建时间': plan.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    '发布时间': plan.published_at.strftime('%Y-%m-%d %H:%M:%S') if plan.published_at else '',
                    '取消时间': plan.cancelled_at.strftime('%Y-%m-%d %H:%M:%S') if plan.cancelled_at else '',
                    '取消原因': plan.cancel_reason or ''
                })
        
        # 如果没有数据，添加一行空数据以保持表头
        if not export_data:
            export_data.append({col: '' for col in [
                '计划名称', '计划类型', '计划状态', '开始日期', '结束日期', '计划描述',
                '经营区域', '区域编码', '门店类型', '类型编码', '目标数量', '完成数量',
                '完成率(%)', '贡献率(%)', '预算金额', '创建人', '创建时间', '发布时间',
                '取消时间', '取消原因'
            ]})
        
        # 创建DataFrame
        df = pd.DataFrame(export_data)
        
        # 生成Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='开店计划数据', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['开店计划数据']
            
            # 设置列宽
            column_widths = {
                'A': 25,  # 计划名称
                'B': 12,  # 计划类型
                'C': 12,  # 计划状态
                'D': 12,  # 开始日期
                'E': 12,  # 结束日期
                'F': 30,  # 计划描述
                'G': 15,  # 经营区域
                'H': 12,  # 区域编码
                'I': 15,  # 门店类型
                'J': 12,  # 类型编码
                'K': 12,  # 目标数量
                'L': 12,  # 完成数量
                'M': 12,  # 完成率
                'N': 12,  # 贡献率
                'O': 15,  # 预算金额
                'P': 15,  # 创建人
                'Q': 20,  # 创建时间
                'R': 20,  # 发布时间
                'S': 20,  # 取消时间
                'T': 30   # 取消原因
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置表头样式
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        output.seek(0)
        return output.getvalue()
    
    def get_export_statistics(self, plan_ids=None, start_date=None, end_date=None, 
                             plan_type=None, status=None) -> Dict[str, Any]:
        """获取导出数据统计信息"""
        # 构建查询条件
        queryset = StorePlan.objects.all()
        
        if plan_ids:
            queryset = queryset.filter(id__in=plan_ids)
        
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        if plan_type:
            queryset = queryset.filter(plan_type=plan_type)
        
        if status:
            queryset = queryset.filter(status=status)
        
        # 统计信息
        total_plans = queryset.count()
        total_regional_plans = RegionalPlan.objects.filter(plan__in=queryset).count()
        
        # 按状态统计
        status_counts = {}
        for status_code, status_name in StorePlan.STATUS_CHOICES:
            count = queryset.filter(status=status_code).count()
            if count > 0:
                status_counts[status_name] = count
        
        # 按类型统计
        type_counts = {}
        for type_code, type_name in StorePlan.PLAN_TYPE_CHOICES:
            count = queryset.filter(plan_type=type_code).count()
            if count > 0:
                type_counts[type_name] = count
        
        return {
            'total_plans': total_plans,
            'total_regional_plans': total_regional_plans,
            'status_distribution': status_counts,
            'type_distribution': type_counts,
            'date_range': {
                'earliest_start': queryset.aggregate(
                    earliest=models.Min('start_date')
                )['earliest'],
                'latest_end': queryset.aggregate(
                    latest=models.Max('end_date')
                )['latest']
            }
        }
    
    def generate_advanced_import_template(self, include_sample_data=True, template_type='standard') -> bytes:
        """生成高级导入模板Excel文件"""
        
        if template_type == 'standard':
            # 标准模板（包含示例数据）
            sample_data = [
                {
                    '计划名称': '2024年华东区开店计划',
                    '计划类型': '年度计划',
                    '开始日期': '2024-01-01',
                    '结束日期': '2024-12-31',
                    '计划描述': '2024年华东区域开店计划',
                    '经营区域': '华东区',
                    '区域编码': 'HD',
                    '门店类型': '直营店',
                    '类型编码': 'ZY',
                    '目标数量': 50,
                    '贡献率(%)': 30.5,
                    '预算金额': 5000000
                },
                {
                    '计划名称': '2024年华东区开店计划',
                    '计划类型': '年度计划',
                    '开始日期': '2024-01-01',
                    '结束日期': '2024-12-31',
                    '计划描述': '2024年华东区域开店计划',
                    '经营区域': '华东区',
                    '区域编码': 'HD',
                    '门店类型': '加盟店',
                    '类型编码': 'JM',
                    '目标数量': 30,
                    '贡献率(%)': 20.0,
                    '预算金额': 3000000
                }
            ] if include_sample_data else [{}]
            
        elif template_type == 'quarterly':
            # 季度计划模板
            sample_data = [
                {
                    '计划名称': '2024年Q1华南区开店计划',
                    '计划类型': '季度计划',
                    '开始日期': '2024-01-01',
                    '结束日期': '2024-03-31',
                    '计划描述': '2024年第一季度华南区域开店计划',
                    '经营区域': '华南区',
                    '区域编码': 'HN',
                    '门店类型': '直营店',
                    '类型编码': 'ZY',
                    '目标数量': 15,
                    '贡献率(%)': 25.0,
                    '预算金额': 1500000
                }
            ] if include_sample_data else [{}]
            
        elif template_type == 'bulk':
            # 批量导入模板（更多示例数据）
            regions = [
                ('华东区', 'HD'), ('华南区', 'HN'), ('华北区', 'HB'), ('华中区', 'HZ')
            ]
            store_types = [
                ('直营店', 'ZY'), ('加盟店', 'JM'), ('旗舰店', 'QJ')
            ]
            
            sample_data = []
            for i, (region_name, region_code) in enumerate(regions):
                for j, (store_type_name, store_type_code) in enumerate(store_types):
                    sample_data.append({
                        '计划名称': f'2024年{region_name}开店计划',
                        '计划类型': '年度计划',
                        '开始日期': '2024-01-01',
                        '结束日期': '2024-12-31',
                        '计划描述': f'2024年{region_name}开店计划',
                        '经营区域': region_name,
                        '区域编码': region_code,
                        '门店类型': store_type_name,
                        '类型编码': store_type_code,
                        '目标数量': 20 + i * 5 + j * 3,
                        '贡献率(%)': 15.0 + i * 2.5 + j * 1.5,
                        '预算金额': (20 + i * 5 + j * 3) * 100000
                    })
            
            if not include_sample_data:
                sample_data = [{}]
        else:
            # 空模板
            sample_data = [{}]
        
        # 确保至少有一行数据以保持表头
        if not sample_data:
            sample_data = [{}]
        
        # 创建DataFrame
        df = pd.DataFrame(sample_data)
        
        # 生成Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 主数据表
            df.to_excel(writer, sheet_name='开店计划数据', index=False)
            
            # 添加说明表
            self._add_instruction_sheet(writer)
            
            # 添加基础数据参考表
            self._add_reference_data_sheet(writer)
            
            # 设置主数据表格式
            self._format_main_data_sheet(writer.sheets['开店计划数据'])
        
        output.seek(0)
        return output.getvalue()
    
    def _add_instruction_sheet(self, writer):
        """添加使用说明表"""
        instructions = [
            ['字段名', '说明', '是否必填', '数据类型', '示例值', '备注'],
            ['计划名称', '开店计划的名称', '是', '文本', '2024年华东区开店计划', '不能与现有计划重复'],
            ['计划类型', '计划类型', '是', '文本', '年度计划', '支持：年度计划、季度计划'],
            ['开始日期', '计划开始日期', '是', '日期', '2024-01-01', '格式：YYYY-MM-DD'],
            ['结束日期', '计划结束日期', '是', '日期', '2024-12-31', '必须晚于开始日期'],
            ['计划描述', '计划的详细描述', '否', '文本', '2024年华东区域开店计划', '可选字段'],
            ['经营区域', '经营区域名称', '是', '文本', '华东区', '必须在系统中存在'],
            ['区域编码', '经营区域编码', '否', '文本', 'HD', '与区域名称至少提供一个'],
            ['门店类型', '门店类型名称', '是', '文本', '直营店', '必须在系统中存在'],
            ['类型编码', '门店类型编码', '否', '文本', 'ZY', '与类型名称至少提供一个'],
            ['目标数量', '目标开店数量', '是', '数字', '50', '必须大于0的整数'],
            ['贡献率(%)', '贡献率百分比', '否', '数字', '30.5', '范围0-100，可以有小数'],
            ['预算金额', '预算金额', '否', '数字', '5000000', '不能为负数'],
            ['', '', '', '', '', ''],
            ['业务规则', '', '', '', '', ''],
            ['1. 同一计划中，相同区域和门店类型的组合不能重复', '', '', '', '', ''],
            ['2. 同一计划的总贡献率不能超过100%', '', '', '', '', ''],
            ['3. 计划名称不能与现有计划重复', '', '', '', '', ''],
            ['4. 结束日期必须晚于开始日期', '', '', '', '', ''],
            ['5. 经营区域和门店类型必须在系统中已存在且启用', '', '', '', '', ''],
            ['', '', '', '', '', ''],
            ['导入提示', '', '', '', '', ''],
            ['1. 建议先下载导入模板，在模板基础上修改数据', '', '', '', '', ''],
            ['2. 如果导入失败，请检查错误信息并修正数据后重新导入', '', '', '', '', ''],
            ['3. 大批量数据建议分批导入，每次不超过1000行', '', '', '', '', ''],
            ['4. 导入前请确保相关的经营区域和门店类型已在系统中创建', '', '', '', '', '']
        ]
        
        instruction_df = pd.DataFrame(instructions)
        instruction_df.to_excel(writer, sheet_name='使用说明', index=False, header=False)
        
        # 设置说明表格式
        worksheet = writer.sheets['使用说明']
        
        # 设置列宽
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 30
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['D'].width = 12
        worksheet.column_dimensions['E'].width = 25
        worksheet.column_dimensions['F'].width = 30
        
        # 设置表头样式
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _add_reference_data_sheet(self, writer):
        """添加基础数据参考表"""
        try:
            # 获取经营区域数据
            regions = BusinessRegion.objects.filter(is_active=True).values_list('name', 'code')
            region_data = [['经营区域', '区域编码']] + list(regions)
            
            # 获取门店类型数据
            store_types = StoreType.objects.filter(is_active=True).values_list('name', 'code')
            store_type_data = [['门店类型', '类型编码']] + list(store_types)
            
            # 创建参考数据表
            max_rows = max(len(region_data), len(store_type_data))
            
            # 补齐数据长度
            while len(region_data) < max_rows:
                region_data.append(['', ''])
            while len(store_type_data) < max_rows:
                store_type_data.append(['', ''])
            
            # 合并数据
            reference_data = []
            for i in range(max_rows):
                reference_data.append([
                    region_data[i][0], region_data[i][1], '',
                    store_type_data[i][0], store_type_data[i][1]
                ])
            
            reference_df = pd.DataFrame(reference_data, columns=[
                '经营区域', '区域编码', '', '门店类型', '类型编码'
            ])
            reference_df.to_excel(writer, sheet_name='基础数据参考', index=False)
            
            # 设置参考表格式
            worksheet = writer.sheets['基础数据参考']
            
            # 设置列宽
            worksheet.column_dimensions['A'].width = 15
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 3
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 12
            
        except Exception as e:
            # 如果获取基础数据失败，创建空的参考表
            empty_data = [['经营区域', '区域编码', '', '门店类型', '类型编码']]
            reference_df = pd.DataFrame(empty_data)
            reference_df.to_excel(writer, sheet_name='基础数据参考', index=False, header=False)
    
    def _format_main_data_sheet(self, worksheet):
        """格式化主数据表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 设置列宽
        column_widths = {
            'A': 25,  # 计划名称
            'B': 12,  # 计划类型
            'C': 12,  # 开始日期
            'D': 12,  # 结束日期
            'E': 30,  # 计划描述
            'F': 15,  # 经营区域
            'G': 12,  # 区域编码
            'H': 15,  # 门店类型
            'I': 12,  # 类型编码
            'J': 12,  # 目标数量
            'K': 12,  # 贡献率
            'L': 15   # 预算金额
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置表头样式
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    def get_template_types(self) -> List[Dict[str, Any]]:
        """获取可用的模板类型"""
        return [
            {
                'type': 'standard',
                'name': '标准模板',
                'description': '包含基本示例数据的标准导入模板',
                'sample_count': 2
            },
            {
                'type': 'quarterly',
                'name': '季度计划模板',
                'description': '专门用于季度计划的导入模板',
                'sample_count': 1
            },
            {
                'type': 'bulk',
                'name': '批量导入模板',
                'description': '包含多个区域和门店类型组合的批量导入模板',
                'sample_count': 12
            },
            {
                'type': 'empty',
                'name': '空白模板',
                'description': '不包含示例数据的空白模板',
                'sample_count': 0
            }
        ]